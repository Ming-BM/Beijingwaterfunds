from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from functools import wraps
import json, io, os
try:
    import psycopg2, psycopg2.extras
    HAS_PG = True
except ImportError:
    HAS_PG = False
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'budget-app-secret-2026')

# ===================== 账号配置（改密码后重启生效） =====================
ACCOUNTS = {
    'user': {'password': '123123',  'role': 'user'},   # 填报账号
    'weixin':    {'password': '123123', 'role': 'admin'},  # 管理员账号
    'demo':  {'password': 'demo2026', 'role': 'user'},  # 作品集访客演示账号，提交不落库
}

def login_required(role=None):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if 'username' not in session:
                return redirect('/login')
            if role == 'admin' and session.get('role') != 'admin':
                return '无权限访问', 403
            return f(*args, **kwargs)
        return wrapped
    return decorator

# ===================== 附表类型定义 =====================
SUB_TYPES = {
    'chem_agent':  {'label':'2.5 其他化学药剂',          'main_field':'s1_2_5',  'cols':[('name','药剂名称'),('amount','需求金额(元)'),('quantity','数量(吨)'),('unit_price','单价(元/吨)')]},
    'prod_test':   {'label':'7 生产检测费',               'main_field':'s1_7',    'cols':[('name','项目名称'),('amount','需求金额(元)'),('remark','备注')]},
    'mfg_xm':      {'label':'制造费用—项修（3.1.1）',      'main_field':'s2_3_1a', 'cols':[('name','项目名称'),('amount','需求金额(元)'),('manager','负责人')]},
    'mfg_bj':      {'label':'制造费用—备品备件（3.1.2）',  'main_field':'s2_3_1b', 'cols':[('name','项目名称'),('amount','需求金额(元)'),('manager','负责人')]},
    'mfg_fl':      {'label':'制造费用—辅料（3.1.3）',      'main_field':'s2_3_1c', 'cols':[('name','项目名称'),('amount','需求金额(元)'),('manager','负责人')]},
    'mfg_xzxl':   {'label':'日常中小修—行政性修理（3.2）', 'main_field':'s2_3_2',  'cols':[('name','项目名称'),('amount','需求金额(元)'),('manager','负责人')]},
    'oh_prod':     {'label':'大修—生产性大修（4.1）',      'main_field':'s2_4_1',  'cols':[('name','项目名称'),('amount','需求金额(元)'),('manager','负责人')]},
    'oh_admin':    {'label':'大修—行政性大修（4.2）',      'main_field':'s2_4_2',  'cols':[('name','项目名称'),('amount','需求金额(元)'),('manager','负责人')]},
    'capital':     {'label':'六、资本性支出',              'main_field':'s6',      'cols':[('name','项目名称'),('amount','需求金额(元)'),('code','编号'),('manager','负责人')]},
    'prev_year':   {'label':'七、以前年度挂账',            'main_field':'s7',      'cols':[('name','项目名称'),('amount','需求金额(元)'),('payee','收款单位')]},
    'other_funds': {'label':'八、其他往来资金',            'main_field':'s8',      'cols':[('name','项目名称'),('amount','需求金额(元)'),('remark','备注')]},
}

# ===================== 主表科目定义 =====================
ITEMS = [
    # 一、生产成本
    {'id':'s1','label':'一、生产成本','level':0,'type':'subtotal','children':['s1_1','s1_2','s1_3','s1_4','s1_5','s1_6','s1_7','s1_8']},
    {'id':'s1_1','label':'1、人工成本','level':1,'type':'subtotal','children':['s1_1_1','s1_1_2','s1_1_3','s1_1_4','s1_1_5','s1_1_6','s1_1_7']},
    {'id':'s1_1_1','label':'1.1、生产人员工资','level':2,'type':'input'},
    {'id':'s1_1_2','label':'1.2、福利费','level':2,'type':'input'},
    {'id':'s1_1_3','label':'1.3、社会保险费','level':2,'type':'subtotal','children':['s1_1_3a','s1_1_3b','s1_1_3c','s1_1_3d','s1_1_3e']},
    {'id':'s1_1_3a','label':'（1）养老保险（含年金）','level':3,'type':'input'},
    {'id':'s1_1_3b','label':'（2）医疗保险（含补充医疗）','level':3,'type':'input'},
    {'id':'s1_1_3c','label':'（3）失业保险','level':3,'type':'input'},
    {'id':'s1_1_3d','label':'（4）工伤保险','level':3,'type':'input'},
    {'id':'s1_1_3e','label':'（5）生育保险','level':3,'type':'input'},
    {'id':'s1_1_4','label':'1.4、住房公积金','level':2,'type':'input'},
    {'id':'s1_1_5','label':'1.5、工会经费','level':2,'type':'input'},
    {'id':'s1_1_6','label':'1.6、职工教育经费','level':2,'type':'input'},
    {'id':'s1_1_7','label':'1.7、外包劳务费','level':2,'type':'input'},
    {'id':'s1_2','label':'2、材料费','level':1,'type':'subtotal','children':['s1_2_1','s1_2_2','s1_2_3','s1_2_4','s1_2_5']},
    {'id':'s1_2_1','label':'2.1、絮凝剂','level':2,'type':'input'},
    {'id':'s1_2_2','label':'2.2、除磷药剂','level':2,'type':'input'},
    {'id':'s1_2_3','label':'2.3、脱硫药剂','level':2,'type':'input'},
    {'id':'s1_2_4','label':'2.4、石灰','level':2,'type':'input'},
    {'id':'s1_2_5','label':'2.5、其他化学药剂','level':2,'type':'input','subType':'chem_agent'},
    {'id':'s1_3','label':'3、动力费','level':1,'type':'subtotal','children':['s1_3_1','s1_3_2']},
    {'id':'s1_3_1','label':'3.1、运行热力费','level':2,'type':'input'},
    {'id':'s1_3_2','label':'3.2、运行电力费','level':2,'type':'input'},
    {'id':'s1_4','label':'4、生产用水费','level':1,'type':'input'},
    {'id':'s1_5','label':'5、外购源水','level':1,'type':'input'},
    {'id':'s1_6','label':'6、泥渣砂外运处置费','level':1,'type':'subtotal','children':['s1_6_1','s1_6_2','s1_6_3']},
    {'id':'s1_6_1','label':'6.1、渣砂运费','level':2,'type':'input'},
    {'id':'s1_6_2','label':'6.2、泥运费','level':2,'type':'input'},
    {'id':'s1_6_3','label':'6.3、泥处置费','level':2,'type':'input'},
    {'id':'s1_7','label':'7、生产检测费','level':1,'type':'input','subType':'prod_test'},
    {'id':'s1_8','label':'8、管网泵站维护费','level':1,'type':'subtotal','children':['s1_8_1','s1_8_2','s1_8_3','s1_8_4','s1_8_5','s1_8_6']},
    {'id':'s1_8_1','label':'8.1、生产性日常中小修','level':2,'type':'subtotal','children':['s1_8_1a','s1_8_1b','s1_8_1c']},
    {'id':'s1_8_1a','label':'（1）项修','level':3,'type':'input'},
    {'id':'s1_8_1b','label':'（2）辅料','level':3,'type':'input'},
    {'id':'s1_8_1c','label':'（3）备件','level':3,'type':'input'},
    {'id':'s1_8_2','label':'8.2、生产性大修理','level':2,'type':'input'},
    {'id':'s1_8_3','label':'8.3、机械租赁费','level':2,'type':'input'},
    {'id':'s1_8_4','label':'8.4、生产用劳务费','level':2,'type':'input'},
    {'id':'s1_8_5','label':'8.5、生产用派遣费','level':2,'type':'input'},
    {'id':'s1_8_6','label':'8.6、生产用汽车费','level':2,'type':'input'},
    # 二、制造费用
    {'id':'s2','label':'二、制造费用','level':0,'type':'subtotal','children':['s2_1','s2_3','s2_4','s2_5','s2_6','s2_7','s2_8','s2_9','s2_10','s2_11','s2_12','s2_13','s2_14','s2_15','s2_16','s2_17','s2_18','s2_19','s2_22','s2_23','s2_24','s2_25']},
    {'id':'s2_1','label':'1、人工成本','level':1,'type':'subtotal','children':['s2_1_1','s2_1_2','s2_1_3','s2_1_4','s2_1_5','s2_1_6']},
    {'id':'s2_1_1','label':'1.1、工资','level':2,'type':'input'},
    {'id':'s2_1_2','label':'1.2、职工福利费','level':2,'type':'input'},
    {'id':'s2_1_3','label':'1.3、社会保险费','level':2,'type':'subtotal','children':['s2_1_3a','s2_1_3b','s2_1_3c','s2_1_3d','s2_1_3e']},
    {'id':'s2_1_3a','label':'（1）养老保险（含年金）','level':3,'type':'input'},
    {'id':'s2_1_3b','label':'（2）医疗保险（含补充医疗）','level':3,'type':'input'},
    {'id':'s2_1_3c','label':'（3）失业保险','level':3,'type':'input'},
    {'id':'s2_1_3d','label':'（4）工伤保险','level':3,'type':'input'},
    {'id':'s2_1_3e','label':'（5）生育保险','level':3,'type':'input'},
    {'id':'s2_1_4','label':'1.4、住房公积金','level':2,'type':'input'},
    {'id':'s2_1_5','label':'1.5、工会经费','level':2,'type':'input'},
    {'id':'s2_1_6','label':'1.6、职工教育经费','level':2,'type':'input'},
    {'id':'s2_2','label':'2、折旧费','level':1,'type':'excluded'},
    {'id':'s2_3','label':'3、日常中小修理费','level':1,'type':'subtotal','children':['s2_3_1','s2_3_2']},
    {'id':'s2_3_1','label':'（1）生产性修理费','level':2,'type':'subtotal','children':['s2_3_1a','s2_3_1b','s2_3_1c']},
    {'id':'s2_3_1a','label':'（1.1）项修','level':3,'type':'input','subType':'mfg_xm'},
    {'id':'s2_3_1b','label':'（1.2）备品备件','level':3,'type':'input','subType':'mfg_bj'},
    {'id':'s2_3_1c','label':'（1.3）辅助材料','level':3,'type':'input','subType':'mfg_fl'},
    {'id':'s2_3_2','label':'（2）行政性修理','level':2,'type':'input','subType':'mfg_xzxl'},
    {'id':'s2_4','label':'4、大修理费','level':1,'type':'subtotal','children':['s2_4_1','s2_4_2']},
    {'id':'s2_4_1','label':'（1）生产性大修费','level':2,'type':'input','subType':'oh_prod'},
    {'id':'s2_4_2','label':'（2）行政性大修费','level':2,'type':'input','subType':'oh_admin'},
    {'id':'s2_5','label':'5、物料消耗费','level':1,'type':'input'},
    {'id':'s2_6','label':'6、低值易耗品','level':1,'type':'input'},
    {'id':'s2_7','label':'7、办公费','level':1,'type':'subtotal','children':['s2_7_1','s2_7_2','s2_7_3','s2_7_4','s2_7_5','s2_7_6','s2_7_7']},
    {'id':'s2_7_1','label':'（1）办公用品','level':2,'type':'input'},
    {'id':'s2_7_2','label':'（2）办公电话费','level':2,'type':'input'},
    {'id':'s2_7_3','label':'（3）印刷复印费','level':2,'type':'input'},
    {'id':'s2_7_4','label':'（4）通讯费','level':2,'type':'input'},
    {'id':'s2_7_5','label':'（5）图书资料费','level':2,'type':'input'},
    {'id':'s2_7_6','label':'（6）邮费','level':2,'type':'input'},
    {'id':'s2_7_7','label':'（7）其他','level':2,'type':'input'},
    {'id':'s2_8','label':'8、差旅费','level':1,'type':'input'},
    {'id':'s2_9','label':'9、劳动保护费','level':1,'type':'input'},
    {'id':'s2_10','label':'10、水费','level':1,'type':'input'},
    {'id':'s2_11','label':'11、电费','level':1,'type':'input'},
    {'id':'s2_12','label':'12、汽车费用','level':1,'type':'subtotal','children':['s2_12_1','s2_12_2','s2_12_3','s2_12_4','s2_12_5']},
    {'id':'s2_12_1','label':'（1）保险费','level':2,'type':'input'},
    {'id':'s2_12_2','label':'（2）油料费','level':2,'type':'input'},
    {'id':'s2_12_3','label':'（3）修理费','level':2,'type':'input'},
    {'id':'s2_12_4','label':'（4）停车过路费','level':2,'type':'input'},
    {'id':'s2_12_5','label':'（5）其他','level':2,'type':'input'},
    {'id':'s2_13','label':'13、交通费','level':1,'type':'input'},
    {'id':'s2_14','label':'14、供暖费','level':1,'type':'subtotal','children':['s2_14_1','s2_14_2','s2_14_3']},
    {'id':'s2_14_1','label':'（1）燃料费','level':2,'type':'input'},
    {'id':'s2_14_2','label':'（2）运行维护费','level':2,'type':'input'},
    {'id':'s2_14_3','label':'（3）职工供暖费','level':2,'type':'input'},
    {'id':'s2_15','label':'15、绿化费','level':1,'type':'input'},
    {'id':'s2_16','label':'16、物业费','level':1,'type':'input'},
    {'id':'s2_17','label':'17、保安费','level':1,'type':'input'},
    {'id':'s2_18','label':'18、保洁费','level':1,'type':'input'},
    {'id':'s2_19','label':'19、行政检测费','level':1,'type':'input'},
    {'id':'s2_20','label':'20、无形资产摊销','level':1,'type':'excluded'},
    {'id':'s2_21','label':'21、长期待摊费用','level':1,'type':'excluded'},
    {'id':'s2_22','label':'22、劳务派遣费','level':1,'type':'input'},
    {'id':'s2_23','label':'23、租赁费','level':1,'type':'subtotal','children':['s2_23_1','s2_23_2']},
    {'id':'s2_23_1','label':'（1）设施租赁费','level':2,'type':'input'},
    {'id':'s2_23_2','label':'（2）设备租赁费','level':2,'type':'input'},
    {'id':'s2_24','label':'24、消防费','level':1,'type':'input'},
    {'id':'s2_25','label':'25、其他支出','level':1,'type':'input'},
    # 三、管理费用
    {'id':'s3','label':'三、管理费用','level':0,'type':'subtotal','children':['s3_1','s3_2','s3_7','s3_8','s3_9','s3_10','s3_11','s3_12','s3_13','s3_14','s3_15','s3_16','s3_17','s3_18','s3_19','s3_20','s3_21','s3_22','s3_23','s3_24','s3_25','s3_26','s3_27','s3_28','s3_29','s3_30','s3_31','s3_32','s3_33','s3_34','s3_35','s3_36','s3_37']},
    {'id':'s3_1','label':'1、人工成本','level':1,'type':'subtotal','children':['s3_1_1','s3_1_2','s3_1_3','s3_1_4','s3_1_5','s3_1_6']},
    {'id':'s3_1_1','label':'1.1、工资','level':2,'type':'input'},
    {'id':'s3_1_2','label':'1.2、职工福利费','level':2,'type':'input'},
    {'id':'s3_1_3','label':'1.3、社会保险费','level':2,'type':'subtotal','children':['s3_1_3a','s3_1_3b','s3_1_3c','s3_1_3d','s3_1_3e']},
    {'id':'s3_1_3a','label':'（1）养老保险（含年金）','level':3,'type':'input'},
    {'id':'s3_1_3b','label':'（2）医疗保险（含补充医疗）','level':3,'type':'input'},
    {'id':'s3_1_3c','label':'（3）失业保险','level':3,'type':'input'},
    {'id':'s3_1_3d','label':'（4）工伤保险','level':3,'type':'input'},
    {'id':'s3_1_3e','label':'（5）生育保险','level':3,'type':'input'},
    {'id':'s3_1_4','label':'1.4、住房公积金','level':2,'type':'input'},
    {'id':'s3_1_5','label':'1.5、工会经费','level':2,'type':'input'},
    {'id':'s3_1_6','label':'1.6、职工教育经费','level':2,'type':'input'},
    {'id':'s3_2','label':'2、保险费','level':1,'type':'input'},
    {'id':'s3_3','label':'3、折旧费','level':1,'type':'excluded'},
    {'id':'s3_4','label':'4、无形资产摊销','level':1,'type':'excluded'},
    {'id':'s3_5','label':'5、长期待摊费用','level':1,'type':'excluded'},
    {'id':'s3_6','label':'6、存货盘亏','level':1,'type':'excluded'},
    {'id':'s3_7','label':'7、业务招待费','level':1,'type':'input'},
    {'id':'s3_8','label':'8、差旅费','level':1,'type':'input'},
    {'id':'s3_9','label':'9、办公费','level':1,'type':'subtotal','children':['s3_9_1','s3_9_2','s3_9_3','s3_9_4','s3_9_5','s3_9_6','s3_9_7']},
    {'id':'s3_9_1','label':'（1）办公用品','level':2,'type':'input'},
    {'id':'s3_9_2','label':'（2）办公电话费','level':2,'type':'input'},
    {'id':'s3_9_3','label':'（3）印刷复印费','level':2,'type':'input'},
    {'id':'s3_9_4','label':'（4）通讯费','level':2,'type':'input'},
    {'id':'s3_9_5','label':'（5）图书资料费','level':2,'type':'input'},
    {'id':'s3_9_6','label':'（6）邮费','level':2,'type':'input'},
    {'id':'s3_9_7','label':'（7）其他','level':2,'type':'input'},
    {'id':'s3_10','label':'10、税费','level':1,'type':'subtotal','children':['s3_10_1','s3_10_2','s3_10_3','s3_10_4','s3_10_5','s3_10_6']},
    {'id':'s3_10_1','label':'（1）印花税','level':2,'type':'input'},
    {'id':'s3_10_2','label':'（2）车船税','level':2,'type':'input'},
    {'id':'s3_10_3','label':'（3）残保费','level':2,'type':'input'},
    {'id':'s3_10_4','label':'（4）房产税','level':2,'type':'input'},
    {'id':'s3_10_5','label':'（5）土地税','level':2,'type':'input'},
    {'id':'s3_10_6','label':'（6）其他税费','level':2,'type':'input'},
    {'id':'s3_11','label':'11、租赁费','level':1,'type':'input'},
    {'id':'s3_12','label':'12、诉讼费','level':1,'type':'input'},
    {'id':'s3_13','label':'13、聘请中介机构费','level':1,'type':'input'},
    {'id':'s3_14','label':'14、咨询费','level':1,'type':'input'},
    {'id':'s3_15','label':'15、研究与开发费','level':1,'type':'input'},
    {'id':'s3_16','label':'16、技术转让费','level':1,'type':'input'},
    {'id':'s3_17','label':'17、董事会费','level':1,'type':'input'},
    {'id':'s3_18','label':'18、排污费','level':1,'type':'input'},
    {'id':'s3_19','label':'19、筹建费用','level':1,'type':'input'},
    {'id':'s3_20','label':'20、办公楼费用','level':1,'type':'subtotal','children':['s3_20_1','s3_20_2','s3_20_3','s3_20_4','s3_20_5','s3_20_6','s3_20_7','s3_20_8','s3_20_9','s3_20_10','s3_20_11']},
    {'id':'s3_20_1','label':'（1）保安费','level':2,'type':'input'},
    {'id':'s3_20_2','label':'（2）保洁费','level':2,'type':'input'},
    {'id':'s3_20_3','label':'（3）办公楼电费','level':2,'type':'input'},
    {'id':'s3_20_4','label':'（4）办公楼水费','level':2,'type':'subtotal','children':['s3_20_4a','s3_20_4b']},
    {'id':'s3_20_4a','label':'（4.1）自来水','level':3,'type':'input'},
    {'id':'s3_20_4b','label':'（4.2）饮用水','level':3,'type':'input'},
    {'id':'s3_20_5','label':'（5）办公楼修理费','level':2,'type':'input'},
    {'id':'s3_20_6','label':'（6）制冷费','level':2,'type':'input'},
    {'id':'s3_20_7','label':'（7）花木租摆费','level':2,'type':'input'},
    {'id':'s3_20_8','label':'（8）垃圾清运费','level':2,'type':'input'},
    {'id':'s3_20_9','label':'（9）供暖费','level':2,'type':'input'},
    {'id':'s3_20_10','label':'（10）收视费','level':2,'type':'input'},
    {'id':'s3_20_11','label':'（11）其他','level':2,'type':'input'},
    {'id':'s3_21','label':'21、会议费','level':1,'type':'input'},
    {'id':'s3_22','label':'22、车辆使用费','level':1,'type':'subtotal','children':['s3_22_1','s3_22_2','s3_22_3','s3_22_4','s3_22_5']},
    {'id':'s3_22_1','label':'（1）保险费','level':2,'type':'input'},
    {'id':'s3_22_2','label':'（2）油料费','level':2,'type':'input'},
    {'id':'s3_22_3','label':'（3）修理费','level':2,'type':'input'},
    {'id':'s3_22_4','label':'（4）停车过路费','level':2,'type':'input'},
    {'id':'s3_22_5','label':'（5）其他','level':2,'type':'input'},
    {'id':'s3_23','label':'23、消防费','level':1,'type':'input'},
    {'id':'s3_24','label':'24、业务宣传费','level':1,'type':'input'},
    {'id':'s3_25','label':'25、劳动保护费','level':1,'type':'input'},
    {'id':'s3_26','label':'26、广告费','level':1,'type':'input'},
    {'id':'s3_27','label':'27、市内交通费','level':1,'type':'input'},
    {'id':'s3_28','label':'28、培训费','level':1,'type':'input'},
    {'id':'s3_29','label':'29、物料费','level':1,'type':'input'},
    {'id':'s3_30','label':'30、职工加班餐费','level':1,'type':'input'},
    {'id':'s3_31','label':'31、辞退福利','level':1,'type':'input'},
    {'id':'s3_32','label':'32、退休人员费用','level':1,'type':'input'},
    {'id':'s3_33','label':'33、劳务费','level':1,'type':'input'},
    {'id':'s3_34','label':'34、劳务派遣费','level':1,'type':'input'},
    {'id':'s3_35','label':'35、低值易耗品','level':1,'type':'input'},
    {'id':'s3_36','label':'36、办公设备修理费','level':1,'type':'input'},
    {'id':'s3_37','label':'37、其他','level':1,'type':'input'},
    # 四、财务费用
    {'id':'s4','label':'四、财务费用','level':0,'type':'subtotal','children':['s4_2','s4_3','s4_4','s4_7']},
    {'id':'s4_1','label':'1、利息收入','level':1,'type':'excluded'},
    {'id':'s4_2','label':'2、借款利息支出','level':1,'type':'subtotal','children':['s4_2_1','s4_2_2']},
    {'id':'s4_2_1','label':'（1）短期借款利息','level':2,'type':'input'},
    {'id':'s4_2_2','label':'（2）长期借款利息','level':2,'type':'input'},
    {'id':'s4_3','label':'3、金融机构手续费','level':1,'type':'input'},
    {'id':'s4_4','label':'4、转贷费','level':1,'type':'input'},
    {'id':'s4_5','label':'5、汇兑损益','level':1,'type':'excluded'},
    {'id':'s4_6','label':'6、未确认融资费用','level':1,'type':'excluded'},
    {'id':'s4_7','label':'7、其他','level':1,'type':'input'},
    # 五、六、七、八、合计
    {'id':'s5','label':'五、销售费用小计','level':0,'type':'input'},
    {'id':'total_cost','label':'成本费用资金需求合计','level':0,'type':'subtotal','children':['s1','s2','s3','s4','s5'],'isTotal':True},
    {'id':'s6','label':'六、资本性支出资金需求','level':0,'type':'input','subType':'capital'},
    {'id':'s7','label':'七、以前年度挂账资金需求','level':0,'type':'input','subType':'prev_year'},
    {'id':'s8','label':'八、其他往来资金','level':0,'type':'input','subType':'other_funds'},
    {'id':'grand_total','label':'总资金需求合计','level':0,'type':'subtotal','children':['total_cost','s6','s7','s8'],'isTotal':True},
]

# ===================== 数据库 =====================
class DBHelper:
    """统一封装 PostgreSQL 和 SQLite，本地无 DATABASE_URL 时自动用 SQLite"""
    def __init__(self):
        url = os.environ.get('DATABASE_URL')
        if url and HAS_PG:
            self._conn = psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)
            self._pg = True
        else:
            import sqlite3
            self._conn = sqlite3.connect('budget.db')
            self._conn.row_factory = sqlite3.Row
            self._pg = False

    def _q(self, sql):
        return sql if self._pg else sql.replace('%s', '?')

    def fetchall(self, sql, params=()):
        cur = self._conn.cursor()
        cur.execute(self._q(sql), params)
        return cur.fetchall()

    def execute(self, sql, params=()):
        cur = self._conn.cursor()
        cur.execute(self._q(sql), params)
        return cur

    def insert(self, sql, params=()):
        if self._pg:
            cur = self._conn.cursor()
            cur.execute(sql + ' RETURNING id', params)
            return cur.fetchone()['id']
        else:
            cur = self._conn.cursor()
            cur.execute(self._q(sql), params)
            return cur.lastrowid

    def commit(self):
        self._conn.commit()

    def close(self):
        self._conn.close()

def get_db():
    return DBHelper()

def init_db():
    db = get_db()
    if db._pg:
        db.execute('''CREATE TABLE IF NOT EXISTS submissions (
            id SERIAL PRIMARY KEY, submitter_name TEXT NOT NULL,
            department TEXT NOT NULL, submit_time TEXT NOT NULL, main_fields TEXT NOT NULL
        )''')
        db.execute('''CREATE TABLE IF NOT EXISTS sub_items (
            id SERIAL PRIMARY KEY, submission_id INTEGER NOT NULL,
            submitter_name TEXT NOT NULL, department TEXT NOT NULL,
            submit_time TEXT NOT NULL, sub_type TEXT NOT NULL, item_data TEXT NOT NULL
        )''')
        db.commit()
        db.close()
    else:
        import sqlite3
        conn = sqlite3.connect('budget.db')
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS submissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                submitter_name TEXT NOT NULL, department TEXT NOT NULL,
                submit_time TEXT NOT NULL, main_fields TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS sub_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                submission_id INTEGER NOT NULL, submitter_name TEXT NOT NULL,
                department TEXT NOT NULL, submit_time TEXT NOT NULL,
                sub_type TEXT NOT NULL, item_data TEXT NOT NULL
            );
        ''')
        conn.commit()
        conn.close()

# ===================== 路由 =====================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        account = ACCOUNTS.get(username)
        if account and account['password'] == password:
            session['username'] = username
            session['role'] = account['role']
            return redirect('/admin' if account['role'] == 'admin' else '/')
        return render_template('login.html', error='用户名或密码错误')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

@app.route('/')
@login_required()
def index():
    return render_template('form.html')

@app.route('/api/items')
@login_required()
def api_items():
    return jsonify(ITEMS)

@app.route('/api/subtypes')
@login_required()
def api_subtypes():
    result = {k: {'label': v['label'], 'main_field': v['main_field'],
                  'cols': v['cols']} for k, v in SUB_TYPES.items()}
    return jsonify(result)

@app.route('/submit', methods=['POST'])
@login_required()
def submit():
    data = request.json
    name = data.get('submitter_name', '').strip()
    dept = data.get('department', '').strip()
    if not name or not dept:
        return jsonify({'success': False, 'msg': '姓名和部门不能为空'}), 400

    if session.get('username') == 'demo':
        # 演示账号：不写入数据库
        return jsonify({'success': True})

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    main_fields = data.get('main_fields', {})
    sub_items_data = data.get('sub_items', [])

    conn = get_db()
    sid = conn.insert(
        'INSERT INTO submissions (submitter_name,department,submit_time,main_fields) VALUES (%s,%s,%s,%s)',
        (name, dept, now, json.dumps(main_fields, ensure_ascii=False))
    )
    for item in sub_items_data:
        conn.execute(
            'INSERT INTO sub_items (submission_id,submitter_name,department,submit_time,sub_type,item_data) VALUES (%s,%s,%s,%s,%s,%s)',
            (sid, name, dept, now, item['type'], json.dumps(item['data'], ensure_ascii=False))
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/admin')
@login_required(role='admin')
def admin():
    conn = get_db()
    rows = conn.fetchall('SELECT id,submitter_name,department,submit_time FROM submissions ORDER BY submit_time DESC')
    conn.close()
    return render_template('admin.html', submissions=rows)

@app.route('/admin/delete/<int:sid>', methods=['POST'])
@login_required(role='admin')
def delete_submission(sid):
    conn = get_db()
    conn.execute('DELETE FROM sub_items WHERE submission_id=%s', (sid,))
    conn.execute('DELETE FROM submissions WHERE id=%s', (sid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ===================== Excel 导出 =====================
def to_float(v):
    try:
        return float(str(v).replace(',', '') or 0)
    except:
        return 0.0

def fmt(v):
    return f'{v:,.2f}' if v else '-'

def compute_items(agg):
    result = {}
    def calc(item_id):
        item = next((i for i in ITEMS if i['id'] == item_id), None)
        if not item:
            return 0
        if item['type'] == 'excluded':
            result[item_id] = 0
        elif item['type'] == 'input':
            result[item_id] = to_float(agg.get(item_id, 0))
        else:
            total = sum(calc(c) for c in item.get('children', []))
            result[item_id] = total
        return result[item_id]
    for item in ITEMS:
        calc(item['id'])
    return result

@app.route('/export/main')
@login_required(role='admin')
def export_main():
    conn = get_db()
    agg = {}
    for row in conn.fetchall('SELECT main_fields FROM submissions'):
        for k, v in json.loads(row['main_fields']).items():
            agg[k] = agg.get(k, 0) + to_float(v)
    for st, info in SUB_TYPES.items():
        total = sum(to_float(json.loads(r['item_data']).get('amount', 0))
                    for r in conn.fetchall('SELECT item_data FROM sub_items WHERE sub_type=%s', (st,)))
        agg[info['main_field']] = total
    conn.close()

    computed = compute_items(agg)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '汇总主表'

    BLUE = '1A3A6B'
    LIGHT_BLUE = 'DCE6F1'
    SUB_BLUE = 'EAF0FB'

    ws.merge_cells('A1:C1')
    ws['A1'] = '2026年7月资金计划汇总主表（所有部门合计）'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:C2')
    ws['A2'] = '单位：元'
    ws['A2'].alignment = Alignment(horizontal='right')

    headers = ['计划项目', '本月申报数（元）', '备注']
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor=BLUE)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for item in ITEMS:
        indent = '　' * item['level']
        label = indent + item['label']
        val = computed.get(item['id'], 0)
        note = '（不计入资金计划）' if item['type'] == 'excluded' else ''
        val_str = '' if item['type'] == 'excluded' else fmt(val)

        ws.append([label, val_str, note])
        r = ws.max_row
        for cell in ws[r]:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        if item.get('isTotal'):
            for cell in ws[r]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(fill_type='solid', fgColor=BLUE)
        elif item['level'] == 0 and item['type'] == 'subtotal':
            for cell in ws[r]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type='solid', fgColor=LIGHT_BLUE)
        elif item['type'] == 'subtotal':
            for cell in ws[r]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type='solid', fgColor=SUB_BLUE)
        elif item['type'] == 'excluded':
            for cell in ws[r]:
                cell.fill = PatternFill(fill_type='solid', fgColor='F5F5F5')
                cell.font = Font(color='AAAAAA')

        ws[r][1].alignment = Alignment(horizontal='right', vertical='center')

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='汇总主表.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/sub')
@login_required(role='admin')
def export_sub():
    conn = get_db()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '汇总附表'

    DARK_BLUE = '1A3A6B'
    HEADER_BLUE = '2E5EA8'
    TOTAL_BG = 'EAF0FB'
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Per-sub-type column definitions
    # keys: None=序号, '_sub'=填报人, '_dept'=所属部门, '_approval'=审批金额(=amount), else data key
    # amt: 1-indexed position of 需求金额 column; apv: position of 审批金额 (None if absent)
    SC = {
        'chem_agent':  {'h':['序号','填报人','所属部门','药剂名称','需求金额(元)','数量(吨)','单价(元/吨)'],
                        'k':[None,'_sub','_dept','name','amount','quantity','unit_price'],'amt':5,'apv':None},
        'prod_test':   {'h':['序号','填报人','所属部门','项目名称','需求金额(元)','备注'],
                        'k':[None,'_sub','_dept','name','amount','remark'],'amt':5,'apv':None},
        'mfg_xm':     {'h':['序号','填报人','所属部门','项目名称','需求金额(元)','审批金额(元)','负责人'],
                        'k':[None,'_sub','_dept','name','amount','_approval','manager'],'amt':5,'apv':6},
        'mfg_bj':     {'h':['序号','填报人','所属部门','项目名称','需求金额(元)','审批金额(元)','负责人'],
                        'k':[None,'_sub','_dept','name','amount','_approval','manager'],'amt':5,'apv':6},
        'mfg_fl':     {'h':['序号','填报人','所属部门','项目名称','需求金额(元)','审批金额(元)','负责人'],
                        'k':[None,'_sub','_dept','name','amount','_approval','manager'],'amt':5,'apv':6},
        'mfg_xzxl':  {'h':['序号','填报人','所属部门','项目名称','需求金额(元)','审批金额(元)','负责人'],
                        'k':[None,'_sub','_dept','name','amount','_approval','manager'],'amt':5,'apv':6},
        'oh_prod':    {'h':['序号','填报人','所属部门','项目名称','需求金额(元)','审批金额(元)','负责人'],
                        'k':[None,'_sub','_dept','name','amount','_approval','manager'],'amt':5,'apv':6},
        'oh_admin':   {'h':['序号','填报人','所属部门','项目名称','需求金额(元)','审批金额(元)','负责人'],
                        'k':[None,'_sub','_dept','name','amount','_approval','manager'],'amt':5,'apv':6},
        'capital':    {'h':['序号','填报人','所属部门','项目名称','需求金额(元)','审批金额(元)','编号','负责人'],
                        'k':[None,'_sub','_dept','name','amount','_approval','code','manager'],'amt':5,'apv':6},
        'prev_year':  {'h':['序号','填报人','所属部门','项目名称','需求金额(元)','审批金额(元)','收款单位'],
                        'k':[None,'_sub','_dept','name','amount','_approval','payee'],'amt':5,'apv':6},
        'other_funds':{'h':['序号','填报人','所属部门','项目名称','需求金额(元)','备注'],
                        'k':[None,'_sub','_dept','name','amount','remark'],'amt':5,'apv':None},
    }

    # Title row
    ws.merge_cells('A1:H1')
    ws['A1'] = '2026年7月资金计划汇总附表'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(fill_type='solid', fgColor=DARK_BLUE)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    ws.merge_cells('A2:H2')
    ws['A2'] = '单位：元'
    ws['A2'].alignment = Alignment(horizontal='right')
    ws.row_dimensions[2].height = 18

    cur_row = 3

    for st, info in SUB_TYPES.items():
        rows = conn.fetchall(
            'SELECT submitter_name,department,item_data FROM sub_items WHERE sub_type=%s ORDER BY id',
            (st,)
        )

        sc = SC[st]
        ncols = len(sc['h'])
        col_end = get_column_letter(ncols)

        # Section title row
        ws.merge_cells(f'A{cur_row}:{col_end}{cur_row}')
        c = ws.cell(cur_row, 1)
        c.value = info['label']
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(fill_type='solid', fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # Header row
        for i, h in enumerate(sc['h'], 1):
            c = ws.cell(cur_row, i)
            c.value = h
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill(fill_type='solid', fgColor=HEADER_BLUE)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border
        ws.row_dimensions[cur_row].height = 20
        cur_row += 1

        # Data rows
        total = 0.0
        for idx, row in enumerate(rows, 1):
            d = json.loads(row['item_data'])
            amt = to_float(d.get('amount', 0))
            total += amt
            for i, key in enumerate(sc['k'], 1):
                if key is None:
                    val = idx
                elif key == '_sub':
                    val = row['submitter_name']
                elif key == '_dept':
                    val = row['department']
                elif key == '_approval':
                    val = amt
                elif key == 'amount':
                    val = amt
                else:
                    val = d.get(key, '')
                c = ws.cell(cur_row, i)
                c.value = val
                c.border = border
                c.alignment = Alignment(vertical='center')
            ws.cell(cur_row, sc['amt']).alignment = Alignment(horizontal='right', vertical='center')
            if sc['apv']:
                ws.cell(cur_row, sc['apv']).alignment = Alignment(horizontal='right', vertical='center')
            cur_row += 1

        # Total row
        for i in range(1, ncols + 1):
            ws.cell(cur_row, i).fill = PatternFill(fill_type='solid', fgColor=TOTAL_BG)
            ws.cell(cur_row, i).border = border
        ws.cell(cur_row, 1).value = '合计'
        ws.cell(cur_row, 1).font = Font(bold=True)
        ws.cell(cur_row, sc['amt']).value = total
        ws.cell(cur_row, sc['amt']).font = Font(bold=True)
        ws.cell(cur_row, sc['amt']).alignment = Alignment(horizontal='right', vertical='center')
        if sc['apv']:
            ws.cell(cur_row, sc['apv']).value = total
            ws.cell(cur_row, sc['apv']).font = Font(bold=True)
            ws.cell(cur_row, sc['apv']).alignment = Alignment(horizontal='right', vertical='center')
        cur_row += 2  # total row + blank separator

    # Fixed column widths (A-H covers all sections)
    for col, w in zip('ABCDEFGH', [6, 10, 16, 32, 15, 15, 20, 10]):
        ws.column_dimensions[col].width = w

    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='汇总附表.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/preview/main')
@login_required(role='admin')
def api_preview_main():
    conn = get_db()
    agg = {}
    for row in conn.fetchall('SELECT main_fields FROM submissions'):
        for k, v in json.loads(row['main_fields']).items():
            agg[k] = agg.get(k, 0) + to_float(v)
    for st, info in SUB_TYPES.items():
        total = sum(to_float(json.loads(r['item_data']).get('amount', 0))
                    for r in conn.fetchall('SELECT item_data FROM sub_items WHERE sub_type=%s', (st,)))
        agg[info['main_field']] = total
    conn.close()
    computed = compute_items(agg)
    return jsonify({'items': ITEMS, 'values': {k: round(v, 2) for k, v in computed.items()}})


APPROVAL_TYPES = {'mfg_xm','mfg_bj','mfg_fl','mfg_xzxl','oh_prod','oh_admin','capital','prev_year'}

@app.route('/api/preview/sub')
@login_required(role='admin')
def api_preview_sub():
    conn = get_db()
    result = {}
    for st, info in SUB_TYPES.items():
        rows = conn.fetchall(
            'SELECT submitter_name,department,item_data FROM sub_items WHERE sub_type=%s ORDER BY id',
            (st,)
        )
        items = []
        total = 0.0
        for r in rows:
            d = json.loads(r['item_data'])
            total += to_float(d.get('amount', 0))
            items.append({'sub': r['submitter_name'], 'dept': r['department'], 'data': d})
        result[st] = {
            'label': info['label'],
            'cols': info['cols'],
            'has_approval': st in APPROVAL_TYPES,
            'items': items,
            'total': round(total, 2)
        }
    conn.close()
    return jsonify(result)

init_db()

if __name__ == '__main__':
    print('服务已启动：http://localhost:5000')
    print('管理页面：http://localhost:5000/admin')
    app.run(host='0.0.0.0', port=5000, debug=False)
